# -*- coding: utf-8 -*-
"""
Created on Fri Nov  9 10:27:53 2018

@author: jah150330
"""

import os
import time
import datetime
import numpy as np
import pandas as pd
import h5py as h5
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.signal import butter, sosfilt
from ast import literal_eval as make_tuple

class proj_data():
    def __init__(self):
        server = get_proj_dir()+'/data'

        wb = load_workbook(filename=os.path.join(server, 'GlasserROIs.xlsx'))
        ws = wb['Sheet1']

        labels = [str(ws['A'+str(x)].value) for x in range(1,361)]

        for x in range(0,180):
            labels[x] = labels[x] + '_L'
        for x in range(180,360):
            labels[x] = labels[x] + '_R'

        self.roiLabels = labels
        database_file = os.path.join(server,'multimodal_HCP.hdf5')
        self.database = h5.File(database_file, 'r+')

        self.bands = {#'BOLD': (.0005, 1/.72/2), #Bandpass range for HCP rs-fMRI
                      'BOLD bandpass': (.01, .1),
                      #'Slow 4': (.02, .06),
                      #'Slow 3': (.06, .2),
                      #'Slow 2': (.2, .5),
                      #'Slow 1': (.5, 1.5),
                      'Delta': (1.5, 4),
                      'Theta': (4, 8),
                      'Alpha': (8, 12),
                      'Beta': (12, 30),
                      'Gamma': (30, 55)}
        self.colors = []
        with open(server + '/rgb_google20c.txt', 'r') as file:
            for line in file:
                line = line.replace('\n', '')
                line = '(%s)' % line
                self.colors.append(make_tuple(line))

    def get_data(self):
        proj_data={}
        proj_data['roiLabels'] = self.roiLabels
        proj_data['database'] = self.database
        proj_data['bands'] = self.bands
        proj_data['colors'] = self.colors

        return proj_data

    @staticmethod
    def get_meg_metadata(bad_meg_subj=['169040', '662551']):
        #turning this into a  function so I don't have to keep copy-pasting
        pdir = get_proj_dir()
        meg_subj_path = pdir + '/data/timeseries_MEG'
        files = sorted(os.listdir(meg_subj_path), key=str.lower)

        meg_subj = sorted(list(set([f.split('_')[0] for f in files])))
        for bad in bad_meg_subj:
            if bad in meg_subj:
                meg_subj.remove(bad)

        uni_sess = set([f.split('_')[-1].replace('.mat', '') for f in files])
        meg_sess = sorted(list(uni_sess))

        return meg_subj, meg_sess

    @staticmethod
    def get_mri_metadata(bad_mri_subj=['104012', '125525',
                                       '151526', '182840',
                                       '200109', '500222']):
        pdir = get_proj_dir()
        path = pdir + '/data/timeseries_rs-fMRI'
        files = os.listdir(path)

        mri_subj = sorted(list(set([f.split('_')[0] for f in files])))
        for subject in bad_mri_subj:
            if subject in mri_subj:
                mri_subj.remove(subject)

        uni_sess = sorted(list(set([f.split('_')[-1] for f in files])))
        mri_sess = [s.replace('.mat', '') for s in uni_sess]

        return mri_subj, mri_sess


def get_proj_dir():
    nas_path = r"\\utdfs01\UTD\Dept\BBSResearch\LabCLINT" #restricted
    server_path = r"\Projects\1Ongoing\Data analysis_Non UTD"
    project_folder = r"\[201801] Three Modalities in One_Jeff"
    return nas_path + server_path + project_folder

def read_database(dset,labels):
    """
    Read data from HDF5 file into a pandas.DataFrame object and include
    ROI labels

    read_data() assumes an HDF5 structured as follows...
        subject = HCP 6-digit subject code (e.g. 100307)
            mode = MEG or rsfMRI
                session = Session1/2/3 (MEG) or LR/RL (rsfMRI)
                    dset = defaults to 'timeseries'
    """
    temp = np.zeros(dset.shape)
    dset.read_direct(temp)

    df = pd.DataFrame(data=temp,columns=labels)

    return df

def super_corr(x, y):
    """
    Correlating massive matrices (can have uneven number of columns)
    Adapted for this project, without memory checks
    If the matrices are too big, expect RAM troubles
    """
    def center_matrix(a):
        mu_a = a.mean(0)
        mean_mat = np.reshape(np.repeat(mu_a, a.shape[0]), a.shape, order="F")
        return np.subtract(a, mean_mat)

    s = x.shape[0]
    if s != y.shape[0]:
        raise ValueError ("x and y must have the same number of observations")

    std_x = x.std(0, ddof=s - 1)
    std_y = y.std(0, ddof=s - 1)

    cov = np.dot(center_matrix(x).T,center_matrix(y))
    return cov/np.dot(std_x[:, np.newaxis], std_y[np.newaxis, :])

def butter_filter(timeseries, fs, cutoffs, btype='band', order=4):
    #Scipy v1.2.0
    nyquist = fs/2
    butter_cut = np.divide(cutoffs, nyquist) #butterworth param (digital)
    sos = butter(order, butter_cut, output='sos', btype=btype)
    return sosfilt(sos, timeseries)


def ctime():
    ts = time.time()
    return datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')


def plot_scree(eigs, pvals=None, alpha=.05, percent=True, kaiser=False, fname=None):
    mpl.rcParams.update(mpl.rcParamsDefault)

    percent_var = (np.multiply(100, eigs)) / np.sum(eigs)
    cumulative_var = np.zeros(shape=[len(percent_var)])
    c = 0
    for i,p in enumerate(percent_var):
        c = c+p
        cumulative_var[i] = c

    fig, ax = plt.subplots(figsize=(10, 10))
    ax.set_title("Scree plot", fontsize='xx-large')
    ax.plot(np.arange(1, len(percent_var)+1), eigs, '-k')
    ax.set_ylim([0, (max(eigs)*1.2)])
    ax.set_ylabel('Eigenvalues', fontsize='xx-large')
    ax.set_xlabel('Factors', fontsize='xx-large')
#    ax.set_xticklabels(fontsize='xx-large')  # TO-DO: make tick labels bigger
    if percent:
        ax2 = ax.twinx()
        ax2.plot(np.arange(1, len(percent_var)+1), percent_var, 'ok')
        ax2.set_ylim(0, max(percent_var)*1.2)
        ax2.set_ylabel('Percentage of variance explained', fontsize='xx-large')

    if pvals is not None and len(pvals) == len(eigs):
        # TO-DO: add p<.05 legend?
        p_check = [i for i, t in enumerate(pvals) if t < alpha]
        eigencheck = [e for i, e in enumerate(eigs) for j in p_check if i == j]
        ax.plot(np.add(p_check, 1), eigencheck, 'ob', markersize=10)

    if kaiser:
        ax.axhline(1, color='k', linestyle=':', linewidth=2)

    if fname:
        fig.savefig(fname, bbox_inches='tight')
    return fig, ax, ax2


def extract_average_power(hdf5_file, sessions, subjects, rois, image_type, bp=True):
    """
    Extract instantaneous power at each timepoint and average across the signal

    Quick function for calculating power data using our hdf5 hierarchy

    Parameters
    ----------
    hdf5_file : str
    Path to hdf5 file

    sessions : list
    List of session names

    subjects : list
    List of subject codes

    rois : list
    List of ROIs

    image_type : str, "MRI" or "MEG"

    bp : bool, default is False
    Flag for applying a .01 - .1 Hz bandpass filter to the signals

    Returns
    -------
    power_data : dict
    Dictionary of power_data
    """
    power_data = {}
    for sess in sessions:
        session_data = []
        for subj in subjects:
            f = h5.File(hdf5_file, 'r')
            if 'MEG' in image_type:
                h_path = subj + '/MEG/' + sess + '/timeseries'  # '/resampled_truncated'
                # h_path = subj + '/MEG/' + sess + '/resampled_truncated'
                data = f.get(h_path).value
                f.close()
                fs = 500

            if 'MRI' in image_type:
                h_path = subj + '/rsfMRI/' + sess + '/timeseries'
                data = f.get(h_path).value
                f.close()
                fs = 1/.72

            if bp:
                cutoffs = [.01, .1]
                timeseries_data = butter_filter(data, fs, cutoffs)
            else:
                timeseries_data = data

            fft_power = np.absolute(np.fft.rfft(timeseries_data, axis=0))**2
            average_power = np.mean(fft_power, axis=0)
            session_data.append(average_power)

        session_df = pd.DataFrame(np.asarray(session_data),
                                  index=subjects,
                                  columns=rois)
        power_data[sess] = session_df

    return power_data


def save_xls(dict_df, path):
    """
    Save a dictionary of dataframes to an excel file, with each dataframe as a seperate page
    """

    writer = pd.ExcelWriter(path)
    for key in list(dict_df):
        dict_df[key].to_excel(writer, '%s' % key)

    writer.save()


def circ_line_corr(ang, line):
    import scipy as sp
    # Correlate periodic data with linear data
    n = len(ang)
    rxs = sp.stats.pearsonr(line, np.sin(ang))
    rxs = rxs[0]
    rxc = sp.stats.pearsonr(line, np.cos(ang))
    rxc = rxc[0]
    rcs = sp.stats.pearsonr(np.sin(ang), np.cos(ang))
    rcs = rcs[0]
    rho = np.sqrt((rxc ** 2 + rxs ** 2 - 2 * rxc * rxs * rcs) / (1 - rcs ** 2))  # r
    # r_2 = rho**2 #r squared
    pval = 1 - sp.stats.chi2.cdf(n * (rho ** 2), 1)
    # standard_error = np.sqrt((1-r_2)/(n-2))

    return rho, pval  # , r_2,standard_error


def cron_alpha(array):
    k = array.shape[1]  # Columns are the groups
    variances_sum = np.sum(np.var(array, axis=0, ddof=1))
    variances_total = np.var(np.sum(array, axis=1), ddof=1)

    return (k / (k-1)) * (1 - (variances_sum / variances_total))
