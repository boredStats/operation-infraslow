# -*- coding: utf-8 -*-
"""
Created on Fri Nov  9 10:27:53 2018

@author: jah150330
"""

import os
import time
import datetime
import numpy as np
import pandas as pd
import h5py as h5
from openpyxl import load_workbook
from scipy.signal import butter, sosfilt

class proj_data():
    def __init__(self):
        server = _get_proj_dir()+'/data'
#        os.chdir(_get_proj_dir()+'/data')
        
        wb = load_workbook(filename=os.path.join(server, 'GlasserROIs.xlsx'))
        ws = wb['Sheet1']
        
        labels = [str(ws['A'+str(x)].value) for x in range(1,361)]
        
        for x in range(0,180):
            labels[x] = labels[x] + '_L'
        for x in range(180,360):
            labels[x] = labels[x] + '_R'
    
        self.roiLabels = labels
        database_file = os.path.join(server,'multimodal_HCP.hdf5')
        self.database = h5.File(database_file, 'r+')
        
        self.bands = {'BOLD': (.0005, 1/.72/2), #Bandpass range for HCP rs-fMRI
                      'Slow 4': (.02, .06),
                      'Slow 3': (.06, .2),
                      'Slow 2': (.2, .5),
                      'Slow 1': (.5, 1.5),
                      'Delta': (1.5, 4),
                      'Theta': (4, 8),
                      'Alpha': (8, 12),
                      'Beta': (12, 30),
                      'Gamma': (30, 55)}
        
    def get_data(self):
        proj_data={}
        proj_data['roiLabels'] = self.roiLabels
        proj_data['database'] = self.database
        proj_data['bands'] = self.bands
        
        return proj_data
    
    @staticmethod
    def get_meg_metadata(bad_meg_subj=['169040', '662551']):
        #turning this into a  function so I don't have to keep copy-pasting
        pdir = _get_proj_dir()
        meg_subj_path = pdir + '/data/timeseries_MEG'
        files = sorted(os.listdir(meg_subj_path), key=str.lower)
        
        meg_subj = sorted(list(set([f.split('_')[0] for f in files])))
        for bad in bad_meg_subj:
            if bad in meg_subj:
                meg_subj.remove(bad)
                
        uni_sess = set([f.split('_')[-1].replace('.mat', '') for f in files])
        meg_sess = sorted(list(uni_sess))
        
        return meg_subj, meg_sess
    
    @staticmethod
    def get_mri_metadata(bad_mri_subj=['104012', '125525', '151526', '182840', '200109', '500222']):
        pdir = _get_proj_dir()
        path = pdir+'/data/timeseries_rs-fMRI'
        filelist = os.listdir(path)
        filelist = sorted(filelist,key=str.lower)
        
        subj_MRI = sorted(list(set([f.split('_')[0] for f in filelist])))
        sess_MRI = sorted(list(set([f.split('_')[-1].replace('.mat', '') for f in filelist])))

        for subject in bad_mri_subj:
            if subject in subj_MRI:
                subj_MRI.remove(subject)
        
        return subj_MRI, sess_MRI
            
def _get_proj_dir():
    #line 1: server path to parent directory
    #line 2: name of the project folder on server
    server_path = r"\\utdfs01\UTD\Dept\BBSResearch\LabCLINT\Projects\1Ongoing\Data analysis_Non UTD"
    project_folder = r"\[201801] Three Modalities in One_Jeff"
    return server_path + project_folder

def read_database(dset,labels):
        """
        Read data from HDF5 file into a pandas.DataFrame object and include
        ROI labels
        
        read_data() assumes an HDF5 structured as follows...
            subject = HCP 6-digit subject code (e.g. 100307)
                mode = MEG or rsfMRI
                    session = Session1/2/3 (MEG) or LR/RL (rsfMRI)
                        dset = defaults to 'timeseries'
        """
        temp = np.zeros(dset.shape)
        dset.read_direct(temp)
        
        df = pd.DataFrame(data=temp,columns=labels)
        
        return df

def super_corr(x, y):
    """
    Correlating massive matrices (can have uneven number of columns)
    Adapted for this project, without memory checks
    If the matrices are too big, expect RAM troubles
    """
    def center_matrix(a):
        mu_a = a.mean(0)
        mean_mat = np.reshape(np.repeat(mu_a, a.shape[0]), a.shape, order="F")
        return np.subtract(a, mean_mat)
        
    s = x.shape[0]    
    if s != y.shape[0]:
        raise ValueError ("x and y must have the same number of observations")
    
    std_x = x.std(0, ddof=s - 1)
    std_y = y.std(0, ddof=s - 1)
    
    cov = np.dot(center_matrix(x).T,center_matrix(y))
    return cov/np.dot(std_x[:, np.newaxis], std_y[np.newaxis, :])
    
def butter_filter(timeseries, fs, cutoffs, order=4):
    #Scipy v1.2.0
    nyq = fs/2
    butter_cut = np.divide(cutoffs, nyq) #butterworth filter param (digital)
    sos = butter(order, butter_cut, output='sos', btype='band')
    return sosfilt(sos, timeseries)

def ctime():
    ts = time.time()
    return datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')

def _get_proj_dir_dep():
    #Note: Depcrecated, server path is now hard coded into proj_utils
    #This version only workers on scripts that are on the server
    pdir = os.path.abspath(os.getcwd())
    target = "this_is_proj_dir.txt"
    while not [t for t in os.listdir(pdir) if target in t]:
        pdir = os.path.abspath(os.path.dirname(pdir))
    return pdir
