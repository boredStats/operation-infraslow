# -*- coding: utf-8 -*-
"""
Created on Thu Nov  8 12:46:30 2018

@author: jah150330
"""
import os
from openpyxl import load_workbook

os.chdir('//utdfs01/UTD/Dept/BBSResearch/LabCLINT/Projects/1Ongoing/Data analysis_Non UTD/[201801] Three Modalities in One_Jeff/data')

wb = load_workbook(filename='GlasserROIs.xlsx')
ws = wb['Sheet1']

labels = [str(ws['A'+str(x)].value) for x in range(1,361)]

for x in range(0,180):
    labels[x] = labels[x] + '_L'
for x in range(180,360):
    labels[x] = labels[x] + '_R'

import scipy.io as sio
import numpy as np
#import pandas as pd
import h5py as h5

database = h5.File('multimodal_HCP.hdf5','a')

path = r'//utdfs01/UTD/Dept/BBSResearch/LabCLINT/Projects/1Ongoing/Data analysis_Non UTD/[201801] Three Modalities in One_Jeff/data/timeseries_rs-fMRI'

filelist = os.listdir(path)
filelist = sorted(filelist,key=str.lower)

subjectList = [filelist[x].split('_') for x in range(0,66960)]
subjectList = subjectList[::720]

dataMatrix = np.zeros([1200,720])
dataLR = np.zeros([1200,360])
dataRL = np.zeros([1200,360])

for x in range(0,93):
    for index in range(0,720):
        subject = filelist[index*(x+1)][0:6]
        session = filelist[index*(x+1)][7:-7]
        
        matFile = sio.loadmat(path+'/'+filelist[index*(x+1)])
        data = matFile['nonzero_'+session].flatten()
        
        dataMatrix[:,index] = data
    
    dataLR = dataMatrix[:,::2]
    dataRL = dataMatrix[:,1::2]
    
    # /SUBJ_CODE/rsfMRI/SESSION/TIMESERIES
    grpLR = database.require_group('/' + subjectList[x][0] + '/rsfMRI/LR')
    dsetLR = grpLR.require_dataset('timeseries', data=dataLR, shape=(1200,360),
                                  dtype='f',chunks=True, compression='gzip')
    grpRL = database.require_group('/' + subjectList[x][0] + '/rsfMRI/RL')
    dsetRL = grpRL.require_dataset('timeseries', data=dataRL, shape=(1200,360),
                                  dtype='f',chunks=True, compression='gzip')

for name in database['169040']:
    print(name)