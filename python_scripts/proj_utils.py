# -*- coding: utf-8 -*-
"""
Created on Fri Nov  9 10:27:53 2018

@author: jah150330
"""

import os
import time
import datetime
import h5py as h5
import numpy as np
import pandas as pd
import pickle as pkl
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.signal import butter, sosfilt
from ast import literal_eval as make_tuple


class proj_data:
    def __init__(self):
        data_path = '../data/'

        wb = load_workbook(filename=os.path.join(data_path, 'GlasserROIs.xlsx'))
        ws = wb['Sheet1']

        labels = [str(ws['A'+str(x)].value) for x in range(1,361)]

        for x in range(0,180):
            labels[x] = labels[x] + '_L'
        for x in range(180,360):
            labels[x] = labels[x] + '_R'

        self.roiLabels = labels
        database_file = os.path.join(data_path, 'multimodal_HCP.hdf5')
        self.database = h5.File(database_file, 'r+')

        self.bands = {#'BOLD': (.0005, 1/.72/2), #Bandpass range for HCP rs-fMRI
                      'BOLD bandpass': (.01, .1),
                      #'Slow 4': (.02, .06),
                      #'Slow 3': (.06, .2),
                      #'Slow 2': (.2, .5),
                      #'Slow 1': (.5, 1.5),
                      'Delta': (1.5, 4),
                      'Theta': (4, 8),
                      'Alpha': (8, 12),
                      'Beta': (12, 30),
                      'Gamma': (30, 55)}
        self.colors = []
        with open(data_path + 'rgb_google20c.txt', 'r') as file:
            for line in file:
                line = line.replace('\n', '')
                line = '(%s)' % line
                self.colors.append(make_tuple(line))

    def get_data(self):
        proj_data={}
        proj_data['roiLabels'] = self.roiLabels
        proj_data['database'] = self.database
        proj_data['bands'] = self.bands
        proj_data['colors'] = self.colors

        return proj_data

    @staticmethod
    def get_meg_metadata(bad_meg_subj=['169040', '662551']):
        with open('../data/proj_metatdata.pkl', 'rb') as file:
            metadata = pkl.load(file)
        meg_subj = metadata['meg_subj']
        meg_sess = metadata['meg_sess']
        for b in bad_meg_subj:
            if b in meg_subj:
                meg_subj.remove(b)

        return meg_subj, meg_sess

    @staticmethod
    def get_mri_metadata(bad_mri_subj=['104012', '125525',
                                       '151526', '182840',
                                       '200109', '500222']):
        with open('../data/proj_metatdata.pkl', 'rb') as file:
            metadata = pkl.load(file)
        mri_subj = metadata['mri_subj']
        mri_sess = metadata['mri_sess']
        for b in bad_mri_subj:
            if b in mri_subj:
                mri_subj.remove(b)

        return mri_subj, mri_sess


def get_proj_dir():
    # Deprecated, project directory now houses code and data
    nas_path = r"\\utdfs01\UTD\Dept\BBSResearch\LabCLINT"   # restricted
    server_path = r"\Projects\1Ongoing\Data analysis_Non UTD"
    project_folder = r"\[201801] Three Modalities in One_Jeff"
    return nas_path + server_path + project_folder


def read_database(dset,labels):
    """
    Read data from HDF5 file into a pandas.DataFrame object and include
    ROI labels

    read_data() assumes an HDF5 structured as follows...
        subject = HCP 6-digit subject code (e.g. 100307)
            mode = MEG or rsfMRI
                session = Session1/2/3 (MEG) or LR/RL (rsfMRI)
                    dset = defaults to 'timeseries'
    """
    temp = np.zeros(dset.shape)
    dset.read_direct(temp)

    df = pd.DataFrame(data=temp,columns=labels)

    return df

def super_corr(x, y):
    """
    Correlating massive matrices (can have uneven number of columns)
    Adapted for this project, without memory checks
    If the matrices are too big, expect RAM troubles
    """
    def center_matrix(a):
        mu_a = a.mean(0)
        mean_mat = np.reshape(np.repeat(mu_a, a.shape[0]), a.shape, order="F")
        return np.subtract(a, mean_mat)

    s = x.shape[0]
    if s != y.shape[0]:
        raise ValueError ("x and y must have the same number of observations")

    std_x = x.std(0, ddof=s - 1)
    std_y = y.std(0, ddof=s - 1)

    cov = np.dot(center_matrix(x).T,center_matrix(y))
    return cov/np.dot(std_x[:, np.newaxis], std_y[np.newaxis, :])


def butter_filter(timeseries, fs, cutoffs, btype='band', order=4):
    # Scipy v1.2.0
    nyquist = fs/2
    butter_cut = np.divide(cutoffs, nyquist)  # butterworth param (digital)
    sos = butter(order, butter_cut, output='sos', btype=btype)
    return sosfilt(sos, timeseries)


def ctime():
    ts = time.time()
    return datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')


def save_xls(dict_df, path):
    """
    Save a dictionary of dataframes to an excel file, with each dataframe as a seperate page
    """

    writer = pd.ExcelWriter(path)
    for key in list(dict_df):
        dict_df[key].to_excel(writer, '%s' % key)

    writer.save()


def circ_line_corr(ang, line):
    import scipy as sp
    # Correlate periodic data with linear data
    n = len(ang)
    rxs = sp.stats.pearsonr(line, np.sin(ang))
    rxs = rxs[0]
    rxc = sp.stats.pearsonr(line, np.cos(ang))
    rxc = rxc[0]
    rcs = sp.stats.pearsonr(np.sin(ang), np.cos(ang))
    rcs = rcs[0]
    rho = np.sqrt((rxc ** 2 + rxs ** 2 - 2 * rxc * rxs * rcs) / (1 - rcs ** 2))  # r
    # r_2 = rho**2 #r squared
    pval = 1 - sp.stats.chi2.cdf(n * (rho ** 2), 1)
    # standard_error = np.sqrt((1-r_2)/(n-2))

    return rho, pval  # , r_2,standard_error


def cron_alpha(array):
    k = array.shape[1]  # Columns are the groups
    variances_sum = np.sum(np.var(array, axis=0, ddof=1))
    variances_total = np.var(np.sum(array, axis=1), ddof=1)

    return (k / (k-1)) * (1 - (variances_sum / variances_total))


def mirror_strfind(strings):
    # Given all possible combinations of strings in a list, find the mirrored strings
    checkdict = {}  # Creating dict for string tests
    for string_1 in strings:
        for string_2 in strings:
            checkdict['%s_%s' % (string_1, string_2)] = False

    yuki, yuno = [], []
    for string_1 in strings:
        for string_2 in strings:
            test = '%s_%s' % (string_1, string_2)
            mir = '%s_%s' % (string_2, string_1)
            if string_1 == string_2:
                checkdict[test] = True
                yuno.append(test)
                continue

            if not checkdict[test] and not checkdict[mir]:
                checkdict[test] = True
                checkdict[mir] = True
                yuki.append(test)
            else:
                yuno.append(test)

    return yuki, yuno


def save_xls(dict_df, path):
    """
    Save a dictionary of dataframes to an excel file, with each dataframe as a seperate page
    """

    writer = pd.ExcelWriter(path)
    for key in list(dict_df):
        dict_df[key].to_excel(writer, '%s' % key)

    writer.save()
